import openpyxl as xl
import sys
from datetime import date
from datetime import timedelta  

raw = xl.load_workbook('raw.xlsx')
raw = raw.active

res_book = xl.Workbook()
res = res_book.create_sheet(title='result.xlsx')
# datas..
id = 0
start_date = 0
end_date = 0
initial_weeks = 0
initial_puts = 0
cumulative_puts = 0
cumulative_weeks = 0

last_date = 0
longest_week = 0

res_row = 1
res_col = 1

# looping
for r in raw.rows:	
	

	# id is changed
	if id != int(r[0].value): 	
		# cum week added
		cumulative_weeks = cumulative_weeks + longest_week
		# print values.. 
		# print id,start_date,end_date,initial_weeks,initial_puts,cumulative_puts,cumulative_weeks 
		res_col = 1
		res.cell(row = res_row, column = res_col, value = id)
		res_col = res_col + 1
		res.cell(row = res_row, column = res_col, value = start_date)
		res_col = res_col + 1
		res.cell(row = res_row, column = res_col, value = end_date)
		res_col = res_col + 1
		res.cell(row = res_row, column = res_col, value = initial_weeks)
		res_col = res_col + 1
		res.cell(row = res_row, column = res_col, value = initial_puts)
		res_col = res_col + 1
		res.cell(row = res_row, column = res_col, value = cumulative_puts)
		res_col = res_col + 1
		res.cell(row = res_row, column = res_col, value = cumulative_weeks)
		res_col = res_col + 1
		
		res_row = res_row + 1
		
		# initialize
		id = int(r[0].value)
		start_date = int(r[1].value)
		initial_weeks = 0
		initial_puts = 0
		cumulative_puts = 0
		cumulative_weeks = 0
		last_date = 0
		longest_week = 0

	
		# three months pass test
	if last_date != 0 and int(r[1].value) != last_date:
		# it's like a whole new person!
		today = date(int(r[1].value)/10000, (int(r[1].value)%10000)/100,  int(r[1].value)%100)
		last =  date(last_date/10000, (last_date%10000)/100,  last_date%100)
		# add longest week
		last_temp = last + timedelta(weeks = (longest_week + 12)) # three month?
		# too far from last drug
		# debug
		#print last, last_temp, today
		if last_temp < today:
			cumulative_weeks = cumulative_weeks + longest_week
			#print id,start_date,end_date,initial_weeks,initial_puts,cumulative_puts,cumulative_weeks 
			res.cell(row = res_row, column = res_col, value = start_date)
			res_col = res_col + 1
			res.cell(row = res_row, column = res_col, value = end_date)
			res_col = res_col + 1
			res.cell(row = res_row, column = res_col, value = initial_weeks)
			res_col = res_col + 1
			res.cell(row = res_row, column = res_col, value = initial_puts)
			res_col = res_col + 1
			res.cell(row = res_row, column = res_col, value = cumulative_puts)
			res_col = res_col + 1
			res.cell(row = res_row, column = res_col, value = cumulative_weeks)
			res_col = res_col + 1
			
			# initialize
			start_date = int(r[1].value)
			initial_weeks = 0
			initial_puts = 0
			cumulative_puts = 0
			cumulative_weeks = 0		
			last_date = 0
			longest_week = 0
			
			
	
	
	
	
	# calculate today's drug
	cumulative_puts = cumulative_puts + r[2].value
		
	# date is changed..
	if int(r[1].value) != last_date:
		# store longest week to cumulative week
		cumulative_weeks = cumulative_weeks + longest_week
		# initialize longest week 
		longest_week = r[3].value
	else:
		# longest week changing
		if longest_week < r[3].value:
			longest_week = r[3].value
	
	# start_date's drug
	if start_date == int(r[1].value):
		initial_puts = initial_puts + r[2].value
		initial_weeks = longest_week
	
	# last update last_date
	last_date = int(r[1].value)
	end_date = int(r[1].value)
	
	
	
cumulative_weeks = cumulative_weeks + longest_week
# print values.. 
#print id,start_date,end_date,initial_weeks,initial_puts,cumulative_puts,cumulative_weeks 
# cum week added
res_col = 1
res.cell(row = res_row, column = res_col, value = id)
res_col = res_col + 1
res.cell(row = res_row, column = res_col, value = start_date)
res_col = res_col + 1
res.cell(row = res_row, column = res_col, value = end_date)
res_col = res_col + 1
res.cell(row = res_row, column = res_col, value = initial_weeks)
res_col = res_col + 1
res.cell(row = res_row, column = res_col, value = initial_puts)
res_col = res_col + 1
res.cell(row = res_row, column = res_col, value = cumulative_puts)
res_col = res_col + 1
res.cell(row = res_row, column = res_col, value = cumulative_weeks)


# save result
res_book.save("result.xlsx")